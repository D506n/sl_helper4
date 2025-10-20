from ._types import modes, table_row
from pathlib import Path
from typing import TextIO
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
import orjson
from ._types import message_callback

class StWildcards():
    start_table:str = '#|'
    end_table:str = '|#'
    start_end_row:str = '||'
    cell_splitter:str = '|'

class FastTable():
    encoding = 'utf-8'
    def __init__(self, cols:list[str], rows:list[table_row]):
        self.cols = cols
        self.rows = rows

    @classmethod
    def read(cls, data:list[table_row]|str|Path|list[str], source_type:modes=None, splitter:str=None, msg_cb:message_callback=None):
        if isinstance(data, str) and Path(data).exists():
            data = Path(data)
        if source_type is None and isinstance(data, Path):
            source_type = data.suffix[1:]
        match source_type:
            case 'json':
                constructor = cls.from_json
            case 'csv':
                constructor = cls.from_csv
            case 'xlsx':
                constructor = cls.from_excel
            case 'startrek':
                constructor = cls.from_startrek
            case _:
                raise AttributeError(f'Неизвестный источник {source_type}')
        return constructor(data, splitter=splitter)

    @classmethod
    def from_json(cls, data:str|Path, *args, **kwargs):
        'Создаёт экземпляр класса RawTable из json. `*args` и `**kwargs` для совместимости с другими конструкторами'
        if isinstance(data, str):
            data:list[table_row] = orjson.loads(data)
        elif isinstance(data, Path):
            file:TextIO = data.open('r', encoding=cls.encoding)
            data = file.read()
            data:list[table_row] = orjson.loads(data)
            file.close()
            if isinstance(data, dict):
                return cls._dict_json_parse(data)
        else:
            raise AttributeError(f'Неизвестный тип данных {type(data)}')
        return cls(cols=list(data[0].keys()), rows=data)

    @classmethod
    def _dict_json_parse(cls, data:dict, *args, **kwargs):
        if 'results' in data.keys():
            try:
                data = data['results'][0]
                cols = []
                for col in data['columns']:
                    cols.append(col['name'])
                items:list[table_row] = []
                for item in data['items']:
                    item:dict[str, table_row]
                    temp = {}
                    for col in cols:
                        temp[col] = item.get(col.lower())
                    items.append(temp)
                return cls(cols=cols, rows=items)
            except Exception:
                raise AttributeError('Неизвестный формат входного json файла.')
        else:
            data = data[list(data.keys())[0]]
            return cls(cols=list(data[0].keys()), rows=data)

    @classmethod
    def from_csv(cls, data:list[str]|str|Path, splitter:str=',', *args, **kwargs):
        'Создаёт экземпляр класса RawTable из csv. `*args` и `**kwargs` для совместимости с другими конструкторами'
        if isinstance(data, Path):
            file:TextIO = data.open('r', encoding=cls.encoding)
            data = file.readlines()
            file.close()
        elif isinstance(data, str):
            data = data.split('\n')
        cols = data.pop(0).replace('\n', '').split(splitter)
        rows:list[table_row] = []
        for row in data:
            rows.append(dict(zip(cols, row.replace('\n', '').split(splitter))))
        return cls(cols=cols, rows=rows)

    @classmethod
    def __from_st_table(cls, st:str):
        step = 0
        headers = []
        rows = []
        raw_rows = st.split(StWildcards.start_end_row)
        while step < len(raw_rows):
            row = raw_rows[step]
            if StWildcards.cell_splitter in row:
                cells = cls.__from_st_row(row)
                if len(headers) == 0:
                    headers = cells
                else:
                    rows.append(cells)
                raw_rows.pop(step)
            else:
                raw_rows.pop(step)
        return cls(cols=headers, rows=rows)

    @classmethod
    def __cell_stripper(cls, cell:str):
        return cell.strip()

    @classmethod
    def __from_st_row(cls, st:str) -> list[str]:
        return list(map(cls.__cell_stripper, st.split(StWildcards.cell_splitter)))

    @classmethod
    def from_startrek(cls, data:str, *args, **kwargs):
        'Создает екземпляр класса RawTable из startrek. `*args` и `**kwargs` для совместимости с другими конструкторами'
        data = data.replace('\n', ' ')
        rows = data.split(StWildcards.start_table)
        step = 0
        while step < len(rows):
            row = rows[step]
            if StWildcards.start_end_row in row:
                table = cls.__from_st_table(row)
                rows.pop(step)
            else:
                rows.pop(step)
        return table

    @classmethod
    def from_excel(cls, data:Path|str, *args, **kwargs):
        'Создаёт екземпляр класса RawTable из excel. `*args` и `**kwargs` для совместимости с другими конструкторами'
        if isinstance(data, str):
            data = cls._parse_excel_str(data)
        elif isinstance(data, Path):
            data = cls._parse_excel_file(data)
        return cls(cols=list(data[0]), rows=data[1:])

    @classmethod
    def _parse_excel_file(cls, path: Path):
        wb = load_workbook(str(path), data_only=True, keep_links=False)
        ws = wb.active
        iterator = ws.iter_rows()
        cols = cls._unpack_excel_row(next(iterator))
        rows:list[table_row] = []
        for vals in iterator:
            rows.append(dict(zip(cols, cls._unpack_excel_row(vals))))
        return rows

    @classmethod
    def _unpack_excel_row(cls, cells: list[Cell]):
        result = []
        for cell in cells:
            result.append(cell.value)
        return result

    @classmethod
    def _parse_excel_str(cls, data: str):
        rows = data.split('\n')
        result:list[table_row] = []
        for row in rows:
            result.append(row.split('\t'))
        return result

    def save(self, dest: Path|None, result_type: modes=None, splitter: str=None, msg_cb: message_callback=None):
        if result_type is None and isinstance(dest, Path):
            result_type = dest.suffix[1:]
        match result_type:
            case 'csv':
                saver = self.to_csv
            case 'json':
                saver = self.to_json
            case 'xlsx':
                saver = self.to_excel
            case 'startrek':
                saver = self.to_startrek
            case _:
                raise AttributeError(f'Неизвестный тип {result_type}')
        return saver(dest, splitter=splitter)

    def json_dumps(self, obj):
        return orjson.dumps(obj, option=orjson.OPT_INDENT_2).decode('utf-8')

    def to_json(self, dest:Path = None, *args, **kwargs) -> str|None:
        if dest is None:
            return self.json_dumps(self.rows)
        with open(dest, 'w', encoding=self.encoding) as file:
            file.write(self.json_dumps(self.rows))

    def to_csv(self, dest:Path = None, splitter:str=',', *args, **kwargs) -> str|None:
        result = []
        result.append(splitter.join(self.cols))
        for row in self.rows:
            temp = list(map(str, row.values()))
            result.append(splitter.join(temp))
        result = '\n'.join(result)
        if dest is None:
            return result
        else:
            file:TextIO = dest.open('w', encoding=self.encoding)
            file.writelines(result)
            file.close()

    @staticmethod
    def _to_excel_cell_str(obj):
        if obj is None:
            return ''
        else:
            return str(obj)

    def _to_string_excel(self):
        result = []
        for row in self.rows:
            result.append('\t'.join(list(map(self.__class__._to_excel_cell_str, row.values()))))
        return '\n'.join(result)

    def to_excel(self, dest:Path=None, *args, **kwargs):
        if dest is None:
            return self._to_string_excel()
        wb = Workbook(True)
        ws = wb.create_sheet('Data')
        ws.append(list(self.cols))
        for row in self.rows:
            ws.append(list(map(self.__class__._to_excel_cell_str, row.values())))
        wb.save(str(dest))

    def to_startrek(self, *args, **kwargs):
        result = [StWildcards.start_table]
        result.append(StWildcards.start_end_row)
        result.append(StWildcards.cell_splitter.join(self.cols))
        result.append(StWildcards.start_end_row)
        for row in self.rows:
            result.append(StWildcards.start_end_row)
            temp = list(map(str, row.values()))
            result.append(StWildcards.cell_splitter.join(temp))
            result.append(StWildcards.start_end_row)
        result.append(StWildcards.end_table)
        return '\n'.join(result)