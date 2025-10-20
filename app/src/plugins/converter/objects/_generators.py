from typing import TextIO
import orjson
from openpyxl import Workbook, load_workbook
from typing import Iterable
from pathlib import Path
from ._types import progress_callback, message_callback

def json_read_gen(file:TextIO, progress:progress_callback=None, msg_cb:message_callback=None):
    result = ''
    square_braces = 0
    curly_braces = 0
    read_progress = 0
    while True:
        line = file.readline()
        if '[' in line and square_braces == 0:
            square_braces += 1
            continue
        if ']' in line and square_braces == 1 and curly_braces == 0:
            break
        if '{' in line and square_braces == 0:
            continue
        if '{' in line and square_braces == 1:
            curly_braces += 1
        if '}' in line:
            curly_braces -= 1
        if '},' in line and curly_braces == 0:
            line = line.replace('},', '}')
        result += line.strip()
        read_progress += 1
        if result.endswith('}'):
            res:dict = orjson.loads(result)
            if progress:
                progress(read_progress)
            yield res
            result = ''

def json_write_gen(file:TextIO, data:Iterable[dict], progress:progress_callback=None, msg_cb:message_callback=None):
    file.write('[\n')
    next_row = next(data)
    last_row = False
    while True:
        yield
        row = next_row
        try:
            next_row = next(data)
        except StopIteration:
            last_row = True
        if not last_row: 
            file.write(f'\t{orjson.dumps(row, option=orjson.OPT_INDENT_2).decode("utf-8")},\n')
        else:
            file.write(f'\t{orjson.dumps(row, option=orjson.OPT_INDENT_2).decode("utf-8")}')
            break
    file.write('\n]')

def csv_read_gen(file:TextIO, splitter:str=',', progress:progress_callback=None, msg_cb:message_callback=None):
    first_row = file.readline()
    headers = first_row.strip().split(splitter)
    read_progress = 0
    while True:
        line = file.readline()
        if line == '':
            break
        row = line.strip().split(splitter)
        yield dict(zip(headers, row))
        if progress is not None:
            read_progress += 1
            progress(read_progress)

def csv_write_gen(file:TextIO, data:Iterable[dict], splitter:str=',', progress:progress_callback=None, msg_cb:message_callback=None):
    first_row = next(data)
    file.write(splitter.join(first_row.keys()) + '\n')
    rows = 0
    while True:
        yield
        try:
            row = next(data)
        except StopIteration:
            break
        file.write(splitter.join([str(row[key]) for key in row.keys()]) + '\n')
        if progress is not None:
            rows += 1
            progress(rows)

def excel_read_gen(file:str|Path, progress:progress_callback=None, msg_cb:message_callback=None):
    wb = load_workbook(file, read_only=True, data_only=True)
    sheet = wb.active
    headers = None
    read_progress = 0
    for row in sheet.iter_rows():
        if headers is None:
            headers = [cell.value for cell in row]
            continue
        result = dict(zip(headers, [cell.value for cell in row]))
        if progress is not None:
            read_progress += 1
            progress(read_progress)
        yield result
    wb.close()

def _to_excel_cell_str(obj):
    if obj is None:
        return ''
    else:
        return str(obj)

def excel_write_gen(file:str|Path, data:Iterable[dict], progress:progress_callback=None, msg_cb:message_callback=None):
    first_row = next(data)
    ws_num = 1
    headers = list(first_row.keys())
    wb = Workbook(True)
    ws = wb.create_sheet(f'Data_{ws_num}')
    ws.append(headers)
    values = list(first_row.values())
    values = list(map(_to_excel_cell_str, values))
    ws.append(values)
    row_num = 1
    for dic in data:
        row_num += 1
        values = list(dic.values())
        values = list(map(_to_excel_cell_str, values))
        ws.append(values)
        if row_num >= 1048570:
            ws_num += 1
            row_num = 1
            ws = wb.create_sheet(f'Data_{ws_num}')
            ws.append(headers)
    try:
        wb.save(file)
    except PermissionError:
        # APP_WINDOW_API.msg.send.emit(f'Не удалось сохранить файл {file}, возможно он занят другим процессом.')
        msg_cb(f'Не удалось сохранить файл {file}, возможно он занят другим процессом.')