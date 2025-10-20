from typing import Generator, TextIO
from pathlib import Path
from ._types import modes, table_row
from openpyxl import load_workbook
from openpyxl.cell import Cell
from ._generators import csv_write_gen, json_write_gen, excel_write_gen, csv_read_gen, json_read_gen, excel_read_gen
from ._types import message_callback, raw_progress_callback
from functools import partial

class OptimizedRows(Generator):
    def __next__(self) -> table_row:
        return super().__next__()

class OptimizedTable():
    encoding = 'utf-8'
    def __init__(self, rows:OptimizedRows, source_file:TextIO=None):
        self.rows = rows
        self.source_file:TextIO = None

    @classmethod
    def read(cls, data:Path, source_type:modes=None, splitter:str=None, msg_cb: message_callback=None, progress: raw_progress_callback=None):
        if isinstance(data, str) and Path(data).exists():
            data = Path(data)
        if source_type is None and isinstance(data, Path):
            source_type = data.suffix[1:]
        match source_type:
            case 'json':
                constructor = cls.from_json
            case 'csv':
                constructor = cls.from_csv
            case 'xlsx':
                constructor = cls.from_excel
            case 'startrek':
                constructor = cls.from_startrek
            case _:
                raise AttributeError(f'Неизвестный источник {source_type}')
        return constructor(data, splitter=splitter, msg_cb=msg_cb, progress=progress)

    @classmethod
    def from_json(cls, data:Path, msg_cb: message_callback=None, progress: raw_progress_callback=None, *args, **kwargs):
        'Создаёт экземпляр класса RawTable из json. `*args` и `**kwargs` для совместимости с другими конструкторами'
        if isinstance(data, Path):
            file:TextIO = data.open('r', encoding=cls.encoding)
            if progress:
                max_count = cls.count_lines(file)
                progress = partial(progress, max_count)
            data:OptimizedRows = json_read_gen(file, progress, msg_cb)
        else:
            raise AttributeError(f'Неизвестный тип данных {type(data)}')
        return cls(rows=data, source_file=file)

    @classmethod
    def from_csv(cls, data:Path, splitter:str=',', msg_cb: message_callback=None, progress: raw_progress_callback=None, *args, **kwargs):
        'Создаёт экземпляр класса RawTable из csv. `*args` и `**kwargs` для совместимости с другими конструкторами'
        if isinstance(data, Path):
            file:TextIO = data.open('r', encoding=cls.encoding)
            if progress:
                max_count = cls.count_lines(file)
                progress = partial(progress, max_count)
            data = csv_read_gen(file, splitter, progress, msg_cb)
        else:
            raise AttributeError(f'Неизвестный тип данных {type(data)}')
        return cls(rows=data, source_file=file)

    @classmethod
    def from_excel(cls, data:Path|str, msg_cb: message_callback=None, progress: raw_progress_callback=None, *args, **kwargs):
        'Создаёт екземпляр класса RawTable из excel. `*args` и `**kwargs` для совместимости с другими конструкторами'
        if isinstance(data, str):
            data = cls._parse_excel_str(data)
        elif isinstance(data, Path):
            max_count = cls.excel_counter(data)
            progress = partial(progress, max_count)
            data = excel_read_gen(data, progress, msg_cb)
        return cls(rows=data)

    @classmethod
    def count_lines(cls, file:TextIO, chunk_size=8192):
        result = sum(chunk.count('\n') for chunk in iter(lambda: file.read(chunk_size), ''))
        file.seek(0)
        return result

    @classmethod
    def excel_counter(cls, file:Path):
        wb = load_workbook(str(file), data_only=True, keep_links=False)
        result = 0
        for ws in wb.worksheets:
            result += ws.max_row
        wb.close()
        return result

    @classmethod
    def from_startrek(cls, data:list[table_row]|str|Path, *args, **kwargs):
        'Создает екземпляр класса RawTable из startrek. `*args` и `**kwargs` для совместимости с другими конструкторами'
        raise NotImplementedError('Появится в будущих обновлениях.')

    @classmethod
    def _unpack_excel_row(cls, cells:list[Cell]):
        result = []
        for cell in cells:
            result.append(cell.value)
        return result

    @classmethod
    def _parse_excel_str(cls, data:str):
        rows = data.split('\n')
        result:list[table_row] = []
        for row in rows:
            result.append(row.split('\t'))
        return result

    def save(self, dest:Path|None, result_type:modes=None, splitter:str=None, msg_cb: message_callback=None):
        if result_type is None and isinstance(dest, Path):
            result_type = dest.suffix[1:]
        match result_type:
            case 'csv':
                saver = self.to_csv
            case 'json':
                saver = self.to_json
            case 'xlsx':
                saver = self.to_excel
            case 'startrek':
                raise NotImplementedError('Возможно появится в будущих обновлениях.')
            case _:
                raise AttributeError(f'Неизвестный тип {result_type}')
        return saver(dest, splitter=splitter, msg_cb=msg_cb)

    def to_json(self, dest:Path, *args, **kwargs) -> str|None:
        with dest.open('w', encoding=self.encoding) as file:
            gen = json_write_gen(file, self.rows)
            for _ in gen:
                pass

    def to_csv(self, dest:Path = None, splitter:str=',', *args, **kwargs) -> str|None:
        file:TextIO = dest.open('w', encoding=self.encoding)
        gen = csv_write_gen(file, self.rows, splitter)
        first_row = next(self.rows)
        file.write(splitter.join(first_row.keys()) + '\n')
        file.write(splitter.join(map(self._to_excel_cell_str, first_row.values())) + '\n')
        for _ in gen:
            pass
        file.close()

    @classmethod
    def _to_excel_cell_str(cls, obj):
        if obj is None:
            return ''
        else:
            return str(obj)

    def to_excel(self, dest:Path=None, *args, **kwargs):
        gen = excel_write_gen(dest, self.rows)
        try:
            for _ in gen:
                pass
        except StopIteration:
            return
        except TypeError:
            return

    def to_startrek(self, *args, **kwargs):
        raise NotImplementedError('Появится в будущих обновлениях.')